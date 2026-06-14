from pathlib import Path
import csv
import sys

import openpyxl


VALID_PLANS = {"A", "G", "G-269", "VIDA", "C"}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Desktop" / "cobranza tres provincias.xlsx"
    target = Path(sys.argv[2]) if len(sys.argv) > 2 else source.with_name("cobranza tres provincias limpio.csv")
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["TICKET"] if "TICKET" in workbook.sheetnames else workbook.worksheets[0]

    rows = [["Poliza", "Plan", "Apellido", "Nombre", "monto cuota", "COBRADOR"]]
    for row in sheet.iter_rows(values_only=True):
        policy = clean(row[1] if len(row) > 1 else "")
        plan = clean(row[2] if len(row) > 2 else "").upper()
        last_name = clean(row[3] if len(row) > 3 else "")
        first_name = clean(row[4] if len(row) > 4 else "")
        amount = clean(row[5] if len(row) > 5 else "")
        collector = clean(row[6] if len(row) > 6 else "").upper() or "OFICINA"
        if policy.isdigit() and plan in VALID_PLANS and (last_name or first_name):
            rows.append([policy, plan, last_name, first_name, amount, collector])

    with target.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerows(rows)

    print(f"CSV generado: {target}")
    print(f"Registros: {len(rows) - 1}")


if __name__ == "__main__":
    main()
